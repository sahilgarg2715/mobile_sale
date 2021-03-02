from openpyxl import load_workbook

wb = load_workbook("my_excel.xlsx")
total_sum = 0
for ws in wb.worksheets:
    for i in range(1, ws.max_row + 1):
        voucher = ws.cell(column=8, row=i).value
        pin = ws.cell(column=9, row=i).value
        amount = ws.cell(column=10, row=i).value
        if i != 1 and voucher is not None and pin is not None:
            total_sum = total_sum + amount

print(total_sum)

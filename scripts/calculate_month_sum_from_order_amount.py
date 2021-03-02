from openpyxl import load_workbook

wb = load_workbook("my_excel.xlsx")
total_sum = 0
for ws in wb.worksheets:
    for i in range(1, ws.max_row + 1):
        amount = str(ws.cell(column=7, row=i).value)
        if i != 1 and amount is not None and amount.isdigit():
            total_sum = total_sum + int(amount)

print(total_sum)

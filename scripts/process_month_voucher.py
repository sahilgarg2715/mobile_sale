from openpyxl import load_workbook
from api.models import MSOrderVouchers

wb = load_workbook("my_excel.xlsx")
MSOrderVouchers.objects.all().delete()
for ws in wb.worksheets:
    for i in range(1, ws.max_row + 1):
        voucher_no = ws.cell(column=8, row=i).value
        pin = str(ws.cell(column=9, row=i).value)
        amount = ws.cell(column=10, row=i).value
        if i != 1 and voucher_no is not None and pin is not None:
            pin = pin.replace("'", "")
            ms_voucher_qs = MSOrderVouchers.objects.filter(
                voucher_no=voucher_no, pin=pin
            )
            if ms_voucher_qs.exists():
                ms_voucher = ms_voucher_qs.first()
                ms_voucher.amount = ms_voucher.amount + amount
                ms_voucher.save()
            else:
                MSOrderVouchers.objects.create(
                    voucher_no=voucher_no, pin=pin, amount=amount
                )

print("done processing")

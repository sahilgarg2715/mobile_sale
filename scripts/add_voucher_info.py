from openpyxl import load_workbook
from api.models import MSReceivedVouchers

wb = load_workbook("voucher_info.xlsx")
MSReceivedVouchers.objects.all().delete()
for ws in wb.worksheets:
    for i in range(1, ws.max_row + 1):
        voucher_no = str(ws.cell(column=1, row=i).value)
        pin = str(ws.cell(column=2, row=i).value)
        amount = ws.cell(column=3, row=i).value
        if voucher_no is not None and pin is not None:
            voucher_no = voucher_no.replace("'", "")
            pin = pin.replace("'", "")
            ms_voucher_info = MSReceivedVouchers.objects.create(
                voucher_no=voucher_no, pin=pin, amount=amount
            )

print("done processing")
